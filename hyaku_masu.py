#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
100マス計算プリント生成プログラム
"""

import random
from datetime import datetime
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("注意: openpyxlがインストールされていません。")
    print("Excel出力を使用するには: pip install openpyxl")


def generate_numbers(count=10, min_val=1, max_val=9):
    """計算用の数字をランダムに生成（1-9を並び替え）"""
    numbers = list(range(min_val, max_val + 1))
    random.shuffle(numbers)
    return numbers[:count]


def print_addition_sheet(row_nums, col_nums):
    """足し算の100マス計算シートを表示"""
    print("\n" + "=" * 60)
    print("【100マス計算 - 足し算】")
    print(f"日付: {datetime.now().strftime('%Y年%m月%d日')}")
    print("=" * 60)
    print("\n問題:")
    
    # ヘッダー行
    print("    +", end="")
    for num in col_nums:
        print(f"{num:4}", end="")
    print()
    print("    " + "-" * (4 * len(col_nums) + 1))
    
    # 各行
    for row_num in row_nums:
        print(f"{row_num:4}|", end="")
        for col_num in col_nums:
            print("    ", end="")  # 空白（答えを書くスペース）
        print()
    
    # 答え
    print("\n" + "=" * 60)
    print("答え:")
    print("    +", end="")
    for num in col_nums:
        print(f"{num:4}", end="")
    print()
    print("    " + "-" * (4 * len(col_nums) + 1))
    
    for row_num in row_nums:
        print(f"{row_num:4}|", end="")
        for col_num in col_nums:
            result = row_num + col_num
            print(f"{result:4}", end="")
        print()
    print("=" * 60)


def print_multiplication_sheet(row_nums, col_nums):
    """掛け算の100マス計算シートを表示"""
    print("\n" + "=" * 60)
    print("【100マス計算 - 掛け算】")
    print(f"日付: {datetime.now().strftime('%Y年%m月%d日')}")
    print("=" * 60)
    print("\n問題:")
    
    # ヘッダー行
    print("    ×", end="")
    for num in col_nums:
        print(f"{num:4}", end="")
    print()
    print("    " + "-" * (4 * len(col_nums) + 1))
    
    # 各行
    for row_num in row_nums:
        print(f"{row_num:4}|", end="")
        for col_num in col_nums:
            print("    ", end="")  # 空白（答えを書くスペース）
        print()
    
    # 答え
    print("\n" + "=" * 60)
    print("答え:")
    print("    ×", end="")
    for num in col_nums:
        print(f"{num:4}", end="")
    print()
    print("    " + "-" * (4 * len(col_nums) + 1))
    
    for row_num in row_nums:
        print(f"{row_num:4}|", end="")
        for col_num in col_nums:
            result = row_num * col_num
            print(f"{result:4}", end="")
        print()
    print("=" * 60)


def save_to_file(filename, calc_type, row_nums, col_nums):
    """計算シートをファイルに保存"""
    with open(filename, 'w', encoding='utf-8') as f:
        f.write("=" * 60 + "\n")
        if calc_type == "add":
            f.write("【100マス計算 - 足し算】\n")
            operator = "+"
        else:
            f.write("【100マス計算 - 掛け算】\n")
            operator = "×"
        
        f.write(f"日付: {datetime.now().strftime('%Y年%m月%d日')}\n")
        f.write("=" * 60 + "\n\n")
        f.write("問題:\n")
        
        # ヘッダー行
        f.write(f"    {operator}")
        for num in col_nums:
            f.write(f"{num:4}")
        f.write("\n")
        f.write("    " + "-" * (4 * len(col_nums) + 1) + "\n")
        
        # 各行
        for row_num in row_nums:
            f.write(f"{row_num:4}|")
            for col_num in col_nums:
                f.write("    ")
            f.write("\n")
        
        # 答え
        f.write("\n" + "=" * 60 + "\n")
        f.write("答え:\n")
        f.write(f"    {operator}")
        for num in col_nums:
            f.write(f"{num:4}")
        f.write("\n")
        f.write("    " + "-" * (4 * len(col_nums) + 1) + "\n")
        
        for row_num in row_nums:
            f.write(f"{row_num:4}|")
            for col_num in col_nums:
                if calc_type == "add":
                    result = row_num + col_num
                else:
                    result = row_num * col_num
                f.write(f"{result:4}")
            f.write("\n")
        f.write("=" * 60 + "\n")


def save_to_excel(filename, calc_type, row_nums, col_nums):
    """計算シートをExcelファイルに保存"""
    if not EXCEL_AVAILABLE:
        print("Error: openpyxlがインストールされていません。")
        return
    
    wb = Workbook()
    
    # 問題シート
    ws_question = wb.active
    ws_question.title = "問題"
    
    # 答えシート
    ws_answer = wb.create_sheet("答え")
    
    if calc_type == "add":
        title = "100マス計算 - 足し算"
        operator = "+"
    else:
        title = "100マス計算 - 掛け算"
        operator = "×"
    
    # スタイル設定
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    answer_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center")
    
    for ws in [ws_question, ws_answer]:
        # タイトル
        ws.merge_cells('A1:L1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center_align
        
        # 日付
        ws.merge_cells('A2:L2')
        ws['A2'] = f"日付: {datetime.now().strftime('%Y年%m月%d日')}"
        ws['A2'].alignment = center_align
        
        # 演算子セル
        ws['B4'] = operator
        ws['B4'].fill = header_fill
        ws['B4'].font = header_font
        ws['B4'].alignment = center_align
        ws['B4'].border = border
        
        # 列ヘッダー（上の数字）
        for i, num in enumerate(col_nums):
            cell = ws.cell(row=4, column=i+3)
            cell.value = num
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = 6
        
        # 行ヘッダー（左の数字）と計算結果
        for i, row_num in enumerate(row_nums):
            # 行ヘッダー
            cell = ws.cell(row=i+5, column=2)
            cell.value = row_num
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            
            # 計算結果
            for j, col_num in enumerate(col_nums):
                cell = ws.cell(row=i+5, column=j+3)
                cell.alignment = center_align
                cell.border = border
                
                if ws == ws_answer:
                    # 答えシートには計算結果を記入
                    if calc_type == "add":
                        cell.value = row_num + col_num
                    else:
                        cell.value = row_num * col_num
                    cell.fill = answer_fill
                # 問題シートは空白のまま
        
        # 行の高さ調整
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[4].height = 20
        for i in range(5, 15):
            ws.row_dimensions[i].height = 20
    
    wb.save(filename)


def main():
    """メイン処理"""
    import os
    
    # outフォルダを作成（存在しない場合）
    os.makedirs('out', exist_ok=True)
    
    print("100マス計算プリント生成プログラム")
    print("-" * 40)
    
    # 計算の種類を選択
    print("\n計算の種類を選択してください:")
    print("1. 足し算")
    print("2. 掛け算")
    print("3. 両方")
    
    choice = input("番号を入力 (1-3): ").strip()
    
    # 数字を生成
    row_nums = generate_numbers(10)
    col_nums = generate_numbers(10)
    
    # 日付文字列を生成
    date_str = datetime.now().strftime('%Y%m%d')
    
    if choice == "1":
        print_addition_sheet(row_nums, col_nums)
        if EXCEL_AVAILABLE:
            filename = f"out/100masu_addition_{date_str}.xlsx"
            save_to_excel(filename, "add", row_nums, col_nums)
            print(f"\n✓ Excelファイルに保存しました: {filename}")
        else:
            filename = f"out/100masu_addition_{date_str}.txt"
            save_to_file(filename, "add", row_nums, col_nums)
            print(f"\n✓ テキストファイルに保存しました: {filename}")
    
    elif choice == "2":
        print_multiplication_sheet(row_nums, col_nums)
        if EXCEL_AVAILABLE:
            filename = f"out/100masu_multiplication_{date_str}.xlsx"
            save_to_excel(filename, "mul", row_nums, col_nums)
            print(f"\n✓ Excelファイルに保存しました: {filename}")
        else:
            filename = f"out/100masu_multiplication_{date_str}.txt"
            save_to_file(filename, "mul", row_nums, col_nums)
            print(f"\n✓ テキストファイルに保存しました: {filename}")
    
    elif choice == "3":
        print_addition_sheet(row_nums, col_nums)
        if EXCEL_AVAILABLE:
            filename = f"out/100masu_addition_{date_str}.xlsx"
            save_to_excel(filename, "add", row_nums, col_nums)
            print(f"\n✓ Excelファイルに保存しました: {filename}")
        else:
            filename = f"out/100masu_addition_{date_str}.txt"
            save_to_file(filename, "add", row_nums, col_nums)
            print(f"\n✓ テキストファイルに保存しました: {filename}")
        
        print_multiplication_sheet(row_nums, col_nums)
        if EXCEL_AVAILABLE:
            filename = f"out/100masu_multiplication_{date_str}.xlsx"
            save_to_excel(filename, "mul", row_nums, col_nums)
            print(f"\n✓ Excelファイルに保存しました: {filename}")
        else:
            filename = f"out/100masu_multiplication_{date_str}.txt"
            save_to_file(filename, "mul", row_nums, col_nums)
            print(f"\n✓ テキストファイルに保存しました: {filename}")
    
    else:
        print("無効な選択です。")


if __name__ == "__main__":
    main()
